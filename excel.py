from openpyxl import load_workbook 
import requests
from openpyxl.drawing.image import Image as ExcelImage
class Excel:
    def __init__(self):
        self.file_path = r"C:\Users\stars\OneDrive\Desktop\shopify_scrapping\products.xlsx"
        self.workbook = load_workbook(filename=self.file_path)
        
    #find the first cell that not filled with value
    def find_last_row(self, sheetname):
        for row in self.workbook[sheetname].iter_rows(max_col=1):
            for cell in row:
                if cell.value is not None:
                    last_cell = cell.coordinate
                    col = last_cell[:1]
                    row = int(last_cell[1:]) + 1
                    
        return(col + str(row))
    
    def fill_in_value(self, sheetname, coordinate, product_info):
        row = coordinate[1:]
        
        self.workbook[sheetname]["A" + str(row)] = product_info["Name"]
        self.workbook[sheetname]["B" + str(row)] = product_info["Price"]
        self.workbook[sheetname]["D" + str(row)] = product_info["Link"]
        
        #insert the product picture
        pic_link = product_info["Picture"]
        
        response = requests.get(pic_link)
        with open(product_info["Name"], "wb") as f:
            f.write(response.content)
        
        img = ExcelImage(product_info["Name"])
        img.width = 100
        img.height = 100
        self.workbook[sheetname].row_dimensions[str(row)].height = 100
        self.workbook[sheetname].column_dimensions["C"].width = 100
        
        coordinate = "C" + str(row)
        
        self.workbook[sheetname].add_image(img, coordinate)
        
        self.workbook.save(self.file_path)
        self.workbook.close()

if __name__ == "__main__":
    excel = Excel()
    print(excel.find_last_row("Dog"))
